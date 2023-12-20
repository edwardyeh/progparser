# SPDX-License-Identifier: GPL-2.0-only
#
# Copyright (C) 2022 Yeh, Hsin-Hsien <yhh76227@gmail.com>
#

"""
Reference table for register parsing
"""

from dataclasses import dataclass, field
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from .general import str2int 


@dataclass (slots=True)
class Reg:
    name:      str
    type:      str
    init_val:  None
    is_access: bool
    addr:      int  = None
    msb:       int  = None
    lsb:       int  = None
    is_signed: bool = None
    comment:   str  = None
    row_idx:   int  = None
    extra:     None = None


@dataclass (slots=True)
class RegList:
    title: str  = None
    regs:  list = field(default_factory=list)


@dataclass (slots=True)
class INIGroup:
    tag:     str
    max_len: int  = 0
    regs:    list = field(default_factory=list)


class ReferenceTable:
    """Reference table for register parsing"""

    def __init__(self, debug_mode: set=None):
        # reg_table = {addr1: reg_list1, addr2: reg_list2, ...}
        # ini_table = [INIGroup1, INIGroup2, ...]
        self.debug_mode = set() if not debug_mode else debug_mode
        self.reg_table = {} 
        self.ini_table = []
        self.hex_out = set()

    def txt_table_parser(self, table_fp: str):
        """Parse text style register table"""
        with open(table_fp, 'r') as f:
            line = f.readline()
            line_no = 1
            while line:
                toks = line.split()
                if len(toks):
                    if toks[0][0] == '#':
                        pass
                    elif toks[0] == 'T:':
                        try:
                            tag_name = ' '.join(toks[1:]).split('#')[0]
                            tag_name = tag_name.strip("\"\' ")
                            self.ini_table.append(INIGroup(tag_name))
                        except Exception as e:
                            print('-' * 60)
                            print("TableParseError: (line: {})".format(line_no))
                            print("syntax of group descriptor:")
                            print("  'T: <tag_name>'")
                            print('-' * 60)
                            raise e
                    elif toks[0] == 'A:':
                        try:
                            addr = str2int(toks[1])
                            reg_list = self.reg_table.setdefault(addr, RegList())
                            if len(toks) > 2:
                                title = ' '.join(toks[2:]).split('#')[0]
                                reg_list.title = title.strip("\"\' ")
                            else:
                                reg_list.title = None
                        except Exception as e:
                            print('-' * 60)
                            print("TableParseError: (line: {})".format(line_no))
                            print("syntax of address descriptor:")
                            print("  'A: <addr> <title>'")
                            print('-' * 60)
                            raise e
                    elif toks[0] == 'H:':
                        for reg in toks[1:]:
                            if reg[0] == '#':
                                break
                            self.hex_out.add(reg.upper())
                    elif toks[0] == '<br>':
                        ini_grp.regs.append(Reg('<br>', *[None]*3))
                    else:
                        reg_name = toks[0].upper()
                        if toks[1] == 'str':
                            try:
                                quote_type = toks[2]
                                init_val = ' '.join(toks[3:])
                                if init_val[0] == '#':
                                    raise e
                                else:
                                    init_val = init_val.split('#')[0]
                                    init_val = init_val.strip("\"\' ")

                                reg = Reg(reg_name, 'str', init_val, True, 
                                          extra=quote_type)
                            except Exception as e:
                                print('-' * 70)
                                print("TableParseError: (line: {})".format(line_no))
                                print("syntax of string descriptor:")
                                print("  '<name> str <init_val> [comment]'")
                                print('-' * 70)
                                raise e
                        elif toks[1] == 'float':
                            try:
                                if len(toks) < 4:
                                    comment = None 
                                else:
                                    comment = self.txt_get_comment(' '.join(toks[3:])) 

                                reg = Reg(reg_name, 'float', float(toks[2]), True, 
                                          comment=comment)
                            except Exception as e:
                                print('-' * 70)
                                print("TableParseError: (line: {})".format(line_no))
                                print("syntax of float descriptor:")
                                print("  '<name> float <init_val> [comment]'")
                                print('-' * 70)
                                raise e
                        elif toks[1] == 'int':
                            try:
                                # msb = str2int(toks[2])
                                # lsb = str2int(toks[3])
                                # is_signed = self.sign_check(toks[4])
                                # init_val = str2int(toks[5], is_signed, msb-lsb+1)
                                if len(toks) < 4:
                                    comment = None
                                else:
                                    comment = self.txt_get_comment(' '.join(toks[3:]))

                                reg = Reg(reg_name, 'int', int(toks[2]), True, 
                                          comment=comment)
                            except Exception as e:
                                print('-' * 70)
                                print("TableParseError: (line: {})".format(line_no))
                                print("syntax of int descriptor:")
                                print("  '<name> int <msb> <lsb> <sign_type> <init_val> [comment]'")
                                print('-' * 70)
                                raise e
                        else:
                            try:
                                addr = str2int(toks[1])
                                msb = str2int(toks[2])
                                lsb = str2int(toks[3])
                                is_signed = self.sign_check(toks[4])
                                is_access = self.access_check(toks[5])
                                init_val = str2int(toks[6], is_signed, msb-lsb+1)
                                if len(toks) < 8:
                                    comment = None
                                else:
                                    comment = self.txt_get_comment(' '.join(toks[7:]))

                                reg = Reg(reg_name, 'reg', init_val, is_access, 
                                          addr=addr, msb=msb, lsb=lsb, 
                                          is_signed=is_signed, comment=comment)

                                reg_list = self.reg_table.setdefault(addr, RegList())
                                reg_list.regs.append(reg)
                            except Exception as e:
                                print('-' * 70)
                                print("TableParseError: (line: {})".format(line_no))
                                print("syntax of register descriptor:")
                                print("  '<name> <addr> <msb> <lsb> <sign_type> <is_access> <init_val> [comment]'")
                                print('-' * 70)
                                raise e

                        if len(self.ini_table):
                            ini_grp = self.ini_table[-1]
                        else:
                            ini_grp = INIGroup(None)
                            self.ini_table.append(ini_grp)

                        reg_len = len(reg_name)
                        if reg_len > ini_grp.max_len:
                            ini_grp.max_len = reg_len
                        ini_grp.regs.append(reg)

                line = f.readline()
                line_no += 1

        if 't' in self.debug_mode:
            self.show_reg_table("=== REG TABLE PARSER ===")
            self.show_ini_table("=== INI TABLE PARSER ===")

    def xlsx_table_parser(self, table_fp: str):
        """Parse excel style reference table"""
        wb = openpyxl.load_workbook(table_fp, data_only=True)
        ws = wb.worksheets[0]
        addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]
        for i in range(addr_col.index('ADDR')+1, len(addr_col)):
            row_idx = i + 1
            if addr_col[i] is not None:
                addr = str(addr_col[i])
                if addr == 'none':
                    break
                else:
                    try:
                        addr = str2int(addr)
                    except Exception as e:
                        print('-' * 60)
                        print("ExcelParseError: (row: {})".format(row_idx))
                        print("address syntax error.")
                        print('-' * 60)
                        raise e

                    if addr in self.reg_table:
                        print('-' * 60)
                        print("ExcelParseError: (row: {})".format(row_idx))
                        print("the addess is existed in the table.")
                        print('-' * 60)
                        raise SyntaxError

                    title = ws.cell(row_idx, 2).value
                    if title is not None:
                        title = str(title).strip()
                    reg_list = RegList(title=title)
                    self.reg_table[addr] = reg_list

            try:
                bits = str(ws.cell(row_idx, 4).value).split('_')
                msb = str2int(bits[0])
                lsb = str2int(bits[1]) if len(bits) > 1 else msb

                try:
                    is_signed = ws.cell(row_idx, 3).font.__getattr__('color').rgb.lower() == 'ff0000ff'
                except Exception:
                    is_signed = False

                init_val = str(ws.cell(row_idx, 3).value)
                if init_val == 'None':
                    init_val = 0 
                else:
                    init_val = str2int(init_val, is_signed, msb-lsb+1) 

                toks = str(ws.cell(row_idx, 5).value).split('\n')
                reg_name = toks[0].strip().upper()
                if len(toks) == 1:
                    comment = None
                else:
                    comment = ', '.join([tok.strip() for tok in toks[1:]])
            except Exception as e:
                print('-' * 60)
                print("ExcelParseError: (row: {})".format(row_idx))
                print("register syntax error (INI/Bits/Member).")
                print('-' * 60)
                raise e

            try:
                is_access = ws.cell(row_idx, 5).font.__getattr__('color').rgb.lower() != 'ff808080'
            except Exception:
                is_access = True

            reg = Reg(reg_name, 'reg', init_val, is_access, addr=addr, 
                      msb=msb, lsb=lsb, is_signed=is_signed,
                      comment=comment, row_idx=row_idx)

            reg_list.regs.append(reg)

            if len(self.ini_table):
                ini_grp = self.ini_table[-1]
            else:
                ini_grp = INIGroup(None)
                self.ini_table.append(ini_grp)

            reg_len = len(reg_name)
            if reg_len > ini_grp.max_len:
                ini_grp.max_len = reg_len
            ini_grp.regs.append(reg)

        wb.close()

        if 't' in self.debug_mode:
            self.show_reg_table("=== XLS TABLE PARSER ===")
            self.show_ini_table("=== INI TABLE PARSER ===")

    def txt_export(self, is_init: bool):
        """Export text style reference table"""
        with open('table_dump.txt', 'w') as f:
            is_first_tag = True
            for ini_grp in self.ini_table:
                if ini_grp.tag is not None:
                    if is_first_tag:
                        is_first_tag = False
                    else:
                        f.write("\n")
                    f.write(f"T: {ini_grp.tag}\n\n")

                max_len = (ini_grp.max_len + 3) >> 2 << 2
                if max_len == ini_grp.max_len:
                    max_len += 4

                for reg in ini_grp.regs:
                    if reg.name == '<br>':
                        f.write("<br>\n")
                        continue

                    if reg.type == 'str':
                        match reg.extra:
                            case 's':
                                reg_val = f"\'{reg.init_val}\'"
                            case 'd':
                                reg_val = f"\"{reg.init_val}\""
                            case _:
                                reg_val = reg.init_val

                        f.write("{}{:<8}{:<4}{}".format(
                            reg.name.lower().ljust(max_len),
                            reg.type,
                            reg.extra,
                            reg_val))
                    elif reg.type == 'float':
                        f.write("{}{:<8}{}".format(
                            reg.name.lower().ljust(max_len),
                            reg.type,
                            reg.init_val))
                    elif reg.type == 'int':
                        f.write("{}{:<8}{:d}".format(
                            reg.name.lower().ljust(max_len),
                            reg.type, 
                            reg.init_val))
                    else:
                        bits = reg.msb - reg.lsb + 1
                        init_val = reg.init_val & ((1 << bits) - 1)
                        f.write("{}{:<#8x}{:<4}{:<4}{:<4}{:<4}{:<#12x}".format(
                            reg.name.lower().ljust(max_len),
                            reg.addr, 
                            reg.msb, 
                            reg.lsb, 
                            's' if reg.is_signed else 'u',
                            'y' if reg.is_access else 'n', 
                            init_val))

                    if reg.comment is not None:
                        f.write(f"\"{reg.comment}\"\n")
                    else:
                        f.write("\n")

            is_first_hex = True
            for reg in self.hex_out:
                if is_first_hex:
                    is_first_hex = False
                    f.write("\n### Hex Out Reg ###\n\n")
                f.write(f"H: {reg.lower()}\n")
            f.write("\n")

            is_first_title = True
            for addr, reg_list in self.reg_table.items():
                if reg_list.title is not None:
                    if is_first_title:
                        is_first_title = False
                        f.write("### Register Title ###\n\n")
                    f.write(f"A: {addr:<#8x}\"{reg_list.title}\"\n")
            f.write("\n")

        if is_init:
            with open('reg_dump.ini', 'w') as f:
                is_first_tag = True
                for ini_grp in self.ini_table:
                    if ini_grp.tag is not None:
                        if is_first_tag:
                            is_first_tag = False
                        else:
                            f.write("\n")
                        f.write(f"[{ini_grp.tag}]\n")

                    for reg in ini_grp.regs:
                        if reg.name == '<br>':
                            f.write("\n")
                            continue

                        if reg.is_access:
                            if reg.name in self.hex_out:
                                if reg.init_val > 0xffff:
                                    lens = ini_grp.max_len + 16
                                    f.write(f"{reg.name.lower()} = {reg.init_val:#010x}".ljust(lens))
                                else:
                                    lens = ini_grp.max_len + 12
                                    f.write(f"{reg.name.lower()} = {reg.init_val:#06x}".ljust(lens))
                            elif reg.type == 'str':
                                match reg.extra:
                                    case 's':
                                        reg_val = f"\'{reg.init_val}\'"
                                    case 'd':
                                        reg_val = f"\"{reg.init_val}\""
                                    case _:
                                        reg_val = reg.init_val

                                lens = ini_grp.max_len + 11
                                f.write(f"{reg.name.lower()} = {reg_val}".ljust(lens))
                            else:
                                lens = ini_grp.max_len + 11
                                f.write(f"{reg.name.lower()} = {reg.init_val}".ljust(lens))

                            if reg.comment is not None:
                                f.write(f" # {reg.comment}\n")
                            else:
                                f.write("\n")

    def xlsx_export(self, is_init: bool, is_rsv_ext: bool=False):
        """Export excel style reference table"""
        GREY_FONT = Font(color='ff808080')
        BLUE_FONT = Font(color='ff0000ff')

        GREEN_FILL = PatternFill(fill_type='solid', start_color='ff92d050')
        GREY_FILL = PatternFill(fill_type='solid', start_color='ffdddddd')
        ORANGE_FILL = PatternFill(fill_type='solid', start_color='ffffcc99')
        YELLOW_FILL = PatternFill(fill_type='solid', start_color='ffffffcc')
        VIOLET_FILL = PatternFill(fill_type='solid', start_color='ffe6ccff')

        THIN_SIDE = Side(border_style='thin', color='ff000000')
        OUTER_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, 
                              top=THIN_SIDE, bottom=THIN_SIDE)

        LT_ALIGN = Alignment(horizontal='left', vertical='top', wrapText=True)
        LC_ALIGN = Alignment(horizontal='left', vertical='center', wrapText=True)
        CT_ALIGN = Alignment(horizontal='center', vertical='top', wrapText=True)
        CC_ALIGN = Alignment(horizontal='center', vertical='center', wrapText=True)
        RT_ALIGN = Alignment(horizontal='right', vertical='top', wrapText=True)
        RC_ALIGN = Alignment(horizontal='right', vertical='center', wrapText=True)

        # Initial workbook 

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        ws.column_dimensions['A'].width = 15.46
        ws.column_dimensions['B'].width = 41.23
        ws.column_dimensions['C'].width = 10.31
        ws.column_dimensions['D'].width = 10.31
        ws.column_dimensions['E'].width = 41.23

        for i in range(1, 3):
            row = ws.row_dimensions[i]
            row.fill = VIOLET_FILL
            row.border = OUTER_BORDER
            row.alignment = CC_ALIGN

        for i in range(3, 7):
            row = ws.row_dimensions[i]
            row.border = OUTER_BORDER
            row.alignment = CC_ALIGN

        row = ws.row_dimensions[7]
        row.fill = GREEN_FILL
        row.border = OUTER_BORDER
        row.alignment = CC_ALIGN

        for row in ws['A1:D6']:
            for cell in row:
                cell.border = OUTER_BORDER
                cell.alignment = CC_ALIGN

        values = ['Chip', 'Eng.', 'Date', 'BaseAddr']
        for row, val in enumerate(values, start=2):
            cell = ws.cell(row, 1, val)
            cell.fill = ORANGE_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

        values = ['Number', 'FileName', 
                  'Metion1', 'Metion2', 'Metion3', 'PatternStatus']
        for row, val in enumerate(values, start=1):
            cell = ws.cell(row, 5, val)
            cell.fill = YELLOW_FILL
            cell.border = OUTER_BORDER
            cell.alignment = LC_ALIGN

        cell = ws.cell(7, 1, 'ADDR')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = CC_ALIGN

        cell = ws.cell(7, 2, 'Register')
        cell.fill = GREEN_FILL 
        cell.border = OUTER_BORDER
        cell.alignment = LC_ALIGN

        cell = ws.cell(7, 3, 'INI')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = RC_ALIGN

        cell = ws.cell(7, 4, 'Bits')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = RC_ALIGN

        cell = ws.cell(7, 5, 'Member')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = LC_ALIGN

        if is_init:
            cell = ws.cell(1, 6, 6)
            cell.fill = VIOLET_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

            cell = ws.cell(2, 6, 'PAT-1')
            cell.fill = VIOLET_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

        # Dump register

        row_st = row_ed = 8
        for addr in range(0, max(self.reg_table.keys()) + 4, 4):
            is_first_reg = True

            if addr in self.reg_table:
                reg_list = self.reg_table[addr]
            elif is_rsv_ext:
                reg_list = RegList(title='reserved', 
                                   regs=[Reg('RESERVED', 'reg', 0, False, 
                                             addr=addr, msb=31, lsb=0, 
                                             is_signed=False)])
            else:
                continue

            for reg in sorted(reg_list.regs, key=lambda reg: reg.lsb):
                cell_fill = YELLOW_FILL if is_first_reg else PatternFill()

                if is_first_reg:
                    is_first_reg = False
                    ws.cell(row_ed, 1, hex(addr))
                    ws.cell(row_ed, 2, reg_list.title)

                cell = ws.cell(row_ed, 1)
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = CT_ALIGN

                cell = ws.cell(row_ed, 2)
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = LT_ALIGN

                init_val = reg.init_val & ((1 << (reg.msb - reg.lsb + 1)) - 1)
                cell = ws.cell(row_ed, 3, hex(init_val))
                cell.font = BLUE_FONT if reg.is_signed else Font()
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = RT_ALIGN

                if reg.msb == reg.lsb:
                    cell = ws.cell(row_ed, 4, str(reg.msb))
                else:
                    cell = ws.cell(row_ed, 4, '_'.join([str(reg.msb), 
                                                        str(reg.lsb)]))
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = RT_ALIGN

                cell_font = Font() if reg.is_access else GREY_FONT

                if reg.name == 'RESERVED':
                    cell = ws.cell(row_ed, 5, reg.name.lower())
                else:
                    toks = [] if reg.comment is None else reg.comment.split(',')
                    members = [reg.name] + [tok.strip() for tok in toks]
                    cell = ws.cell(row_ed, 5, '\n'.join(members))
                cell.font = cell_font
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = LT_ALIGN

                if is_init:
                    cell = ws.cell(row_ed, 6, f"{init_val:X}")
                    cell.font = cell_font
                    cell.fill = cell_fill
                    cell.border = OUTER_BORDER
                    cell.alignment = CC_ALIGN

                row = ws.row_dimensions[row_ed]
                row.font = cell_font
                row.fill = cell_fill
                row.border = OUTER_BORDER
                row.alignment = CC_ALIGN
                row_ed += 1

        row = ws.row_dimensions[row_ed]
        row.fill = GREY_FILL
        row.border = OUTER_BORDER
        row.alignment = CC_ALIGN

        cell = ws.cell(row_ed, 1, 'none')
        cell.fill = GREY_FILL
        cell.border = OUTER_BORDER
        cell.alignment = CC_ALIGN
        row_ed += 2

        cell = ws.cell(row_ed, 1, "Blue INI: ")
        cell.font = BLUE_FONT
        cell.alignment = RC_ALIGN
        cell = ws.cell(row_ed, 2, "signed register")
        cell.font = BLUE_FONT
        cell.alignment = LC_ALIGN
        row_ed += 1

        cell = ws.cell(row_ed, 1, "Grey Member: ")
        cell.font = GREY_FONT
        cell.alignment = RC_ALIGN
        cell = ws.cell(row_ed, 2, "private register")
        cell.font = GREY_FONT
        cell.alignment = LC_ALIGN
        row_ed += 1

        wb.save('table_dump.xlsx')
        wb.close()

    def sign_check(self, str_: str) -> bool:
        """Syntax check for sign type flag"""
        str_ = str_.lower()

        if str_ == 's':
            return True
        elif str_ == 'u':
            return False
        else:
            raise SyntaxError("sign type flag must be 's' or 'u'")

    def access_check(self, str_: str) -> bool:
        """Syntax check for access flag"""
        str_ = str_.lower()

        if str_ == 'y':
            return True
        elif str_ == 'n':
            return False
        else:
            raise SyntaxError("access flag must be 'y' or 'n'")

    def txt_get_comment(self, str_: str) -> str:
        """Get register comment"""
        return None if str_[0] == '#' else str_.split('#')[0].strip("\"\' ")

    def show_reg_table(self, msg: str):
        """Show register table"""
        print(msg, '\n')
        reg_cnt = 0
        for addr, reg_list in self.reg_table.items():
            print(f"Addr: {addr:#06x} Title: {reg_list.title}")
            for reg in reg_list.regs:
                print("  Reg:      {}, {}, {}, {}, {}, {}".format(
                    reg.name, reg.msb, reg.lsb, reg.is_signed, reg.is_access, reg.init_val))
                print("  Comment:  {}".format(reg.comment))
                print("  RowIndex: {}\n".format(reg.row_idx))
                reg_cnt += 1

        print(f"register count: {reg_cnt}\n")

    def show_ini_table(self, msg: str):
        """Show ini table"""
        print(msg, '\n')
        reg_cnt = 0
        for ini_grp in self.ini_table:
            print(f"Tag: {ini_grp.tag}")
            for reg in ini_grp.regs:
                if reg.name == '<br>':
                    print("<br>")
                else:
                    print("  {} = {}".format(
                            reg.name.ljust(ini_grp.max_len),
                            reg.init_val))
                    reg_cnt += 1

        print(f"\nregister count: {reg_cnt}\n")



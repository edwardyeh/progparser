#!/usr/bin/env python3
# -*- coding: utf-8 -*-
## +FHDR=======================================================================
## Copyright (c) 2022 Hsin-Hsien Yeh (Edward Yeh).
## All rights reserved.
## ----------------------------------------------------------------------------
## Filename         : tabconv.py
## File Description : Programming reference table convertor
## ----------------------------------------------------------------------------
## Author           : Edward Yeh
## Created On       : Wed 12 Jan 2022 02:34:24 AM CST
## Format           : Python module
## ----------------------------------------------------------------------------
## Reuse Issues     : 
## ----------------------------------------------------------------------------
## Release History  : 
## -FHDR=======================================================================

import os
import shutil
import argparse
import textwrap
from typing import NamedTuple
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

### Class Definition ###

class Reg(NamedTuple):
##{{{
    name: str
    addr: int
    msb: int
    lsb: int
    is_access: bool
    init_val: int
    comment: str
    row_idx: int
##}}}

@dataclass
class RegList:
#{{{
    title: str
    regs: list
#}}}

@dataclass
class INIGroup:
#{{{ 
    tag: str
    max_len: int
    regs: list
#}}}

class RegisterTable:
    """Programming register table"""

    def __init__(self, table_fp: str, table_type: str, is_debug: bool):
    #{{{ 
        # reg_table = {addr1: reg_list1, addr2: reg_list2, ...}
        # ini_table = [INIGroup1, INIGroup2, ...]

        self.is_debug = is_debug
        self.comment_sign = '#'
        self.reg_table = {} 
        self.ini_table = []

        if table_type == 'txt':
            self.txt_table_parser(table_fp)
        elif table_type == 'xls':
            self.xls_table_parser(table_fp)
        else:
            raise ValueError(f"Unsupported input table style '{table_type}'")
    #}}}

    def txt_table_parser(self, table_fp: str):
        """Parse text style register table"""  #{{{
        with open(table_fp, 'r') as f:
            line = f.readline()
            line_no = 1

            while line:
                toks = line.split()

                if len(toks):
                    if toks[0] == 'T:':
                        try:
                            tag_name = ' '.join(toks[1:]).strip("\"\'")
                            self.ini_table.append(INIGroup(tag_name, 0, []))
                        except Exception as e:
                            print("-" * 60)
                            print("TableParseError: (line: {})".format(line_no))
                            print("syntax of group descriptor:")
                            print("  'T: <tag_name>'")
                            print("-" * 60)
                            raise e

                    elif toks[0] == 'A:':
                        try:
                            addr = self.str2int(toks[1])
                            reg_list = self.reg_table.setdefault(addr, RegList(None, []))
                            reg_list.title = ' '.join(toks[2:]).strip("\"\'") if len(toks) > 2 else None
                        except Exception as e:
                            print("-" * 60)
                            print("TableParseError: (line: {})".format(line_no))
                            print("syntax of address descriptor:")
                            print("  'A: <addr> <title>'")
                            print("-" * 60)
                            raise e

                    else:
                        try:
                            reg_name = toks[0].upper()
                            addr = self.str2int(toks[1])
                            msb = self.str2int(toks[2])
                            lsb = self.str2int(toks[3])
                            is_access = self.access_check(toks[4])
                            init_val = self.str2int(toks[5])
                            comment = ' '.join(toks[6:]).strip("\"\'") if len(toks) > 6 else None
                        except Exception as e:
                            print("-" * 70)
                            print("TableParseError: (line: {})".format(line_no))
                            print("syntax of register descriptor:")
                            print("  '<name> <addr> <msb> <lsb> <is_access> <init_val> [comment]'")
                            print("-" * 70)
                            raise e

                        reg = Reg(reg_name, addr, msb, lsb, is_access, init_val, comment, None)
                        reg_list = self.reg_table.setdefault(addr, RegList(None, []))
                        reg_list.regs.append(reg)

                        if len(self.ini_table):
                            ini_grp = self.ini_table[-1]
                        else:
                            ini_grp = INIGroup(None, 0, [])
                            self.ini_table.append(ini_grp)

                        reg_len = len(reg_name)
                        if reg_len > ini_grp.max_len:
                            ini_grp.max_len = reg_len

                        ini_grp.regs.append(reg)

                line = f.readline()
                line_no += 1

        if self.is_debug:
            self.show_reg_table("=== REG TABLE PARSER ===")
            self.show_ini_table("=== INI TABLE PARSER ===")
    #}}}

    def xls_table_parser(self, table_fp: str):
        """Parse excel style reference table"""  #{{{
        wb = openpyxl.load_workbook(table_fp, data_only=True)
        ws = wb.worksheets[0]
        addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]

        for i in range(addr_col.index('ADDR')+1, len(addr_col)):
            row_idx = i + 1

            if addr_col[i] is not None:
                addr = str(addr_col[i])

                if addr == 'none':
                    break
                else:
                    try:
                        addr = self.str2int(addr)
                    except Exception as e:
                        print("-" * 60)
                        print("ExcelParseError: (row: {})".format(row_idx))
                        print("address syntax error.")
                        print("-" * 60)
                        raise e

                    if addr in self.reg_table:
                        print("-" * 60)
                        print("ExcelParseError: (row: {})".format(row_idx))
                        print("the addess is existed in the table.")
                        print("-" * 60)
                        raise SyntaxError

                    title = ws.cell(row_idx, 2).value
                    if title is not None:
                        title = str(title).strip()
                    reg_list = RegList(title, [])
                    self.reg_table[addr] = reg_list

            try:
                init_val = self.str2int(str(ws.cell(row_idx, 3).value))

                bits = str(ws.cell(row_idx, 4).value).split('_')
                msb = self.str2int(bits[0])
                lsb = self.str2int(bits[1]) if len(bits) > 1 else msb

                toks = str(ws.cell(row_idx, 5).value).split('\n');
                reg_name = toks[0].strip().upper()
                comment = None if len(toks) == 1 else ', '.join([tok.strip() for tok in toks[1:]])
            except Exception as e:
                print("-" * 60)
                print("ExcelParseError: (row: {})".format(row_idx))
                print("register syntax error (INI/Bits/Member).")
                print("-" * 60)
                raise e

            is_access = not ws.cell(row_idx, 5).font.__getattr__('color')
            reg = Reg(reg_name, addr, msb, lsb, is_access, init_val, comment, row_idx)
            reg_list.regs.append(reg)

            if len(self.ini_table):
                ini_grp = self.ini_table[-1]
            else:
                ini_grp = INIGroup(None, 0, [])
                self.ini_table.append(ini_grp)

            reg_len = len(reg_name)
            if reg_len > ini_grp.max_len:
                ini_grp.max_len = reg_len
            ini_grp.regs.append(reg)

        wb.close()

        if self.is_debug:
            self.show_reg_table("=== XLS TABLE PARSER ===")
            self.show_ini_table("=== INI TABLE PARSER ===")
    #}}}

    def txt_export(self, is_init: bool):
        """Export text style reference table"""  #{{{
        with open('table_dump.txt', 'w') as f:
            is_first_tag = True

            for ini_grp in self.ini_table:
                if ini_grp.tag is not None:
                    if is_first_tag:
                        is_first_tag = False
                    else:
                        f.write("\n")

                    f.write(f"T: {ini_grp.tag}\n\n")

                max_len = (ini_grp.max_len + 3) >> 2 << 2

                if max_len == ini_grp.max_len:
                    max_len += 4

                for reg in ini_grp.regs:
                    f.write("{}{}{:<#8x}{:<4}{:<4}{:<4}{:<#12x}".format(
                        reg.name.lower(), " " * (max_len - len(reg.name)),
                        reg.addr, 
                        reg.msb, 
                        reg.lsb, 
                        "y" if reg.is_access else "n", 
                        reg.init_val))

                    if reg.comment is not None:
                        f.write(f"\"{reg.comment}\"\n")
                    else:
                        f.write("\n")

            is_first_title = True

            for addr, reg_list in self.reg_table.items():
                if reg_list.title is not None:
                    if is_first_title:
                        is_first_title = False
                        f.write("\n")
                    f.write(f"A: {addr:<#8x}\"{reg_list.title}\"\n")
            f.write("\n")

        if is_init:
            with open('reg_dump.ini', 'w') as f:
                is_first_tag = True

                for ini_grp in self.ini_table:
                    if ini_grp.tag is not None:
                        if is_first_tag:
                            is_first_tag = False
                        else:
                            f.write("\n")

                        f.write(f"[{ini_grp.tag}]\n")

                    for reg in ini_grp.regs:
                        if reg.is_access:
                            f.write("{}{}".format(reg.name.lower(), " " * (ini_grp.max_len - len(reg.name))))

                            if reg.name.find("ADDR") == -1:
                                f.write(f" = {reg.init_val}")
                            else:
                                f.write(f" = {reg.init_val:<#010x}")

                            if reg.comment is not None:
                                f.write(f"  # {reg.comment}\n")
                            else:
                                f.write("\n")
    #}}}

    def xls_export(self, is_init: bool):
        """Export excel style reference table"""  #{{{
        GREY_FONT = Font(color='808080')

        GREEN_FILL = PatternFill(fill_type='solid', start_color='92d050')
        GREY_FILL = PatternFill(fill_type='solid', start_color='dddddd')
        ORANGE_FILL = PatternFill(fill_type='solid', start_color='ffcc99')
        YELLOW_FILL = PatternFill(fill_type='solid', start_color='ffffcc')
        VIOLET_FILL = PatternFill(fill_type='solid', start_color='e6ccff')

        THIN_SIDE = Side(border_style='thin', color='000000')
        OUTER_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

        LT_ALIGN = Alignment(horizontal='left', vertical='top', wrapText=True)
        LC_ALIGN = Alignment(horizontal='left', vertical='center', wrapText=True)
        CT_ALIGN = Alignment(horizontal='center', vertical='top', wrapText=True)
        CC_ALIGN = Alignment(horizontal='center', vertical='center', wrapText=True)
        RT_ALIGN = Alignment(horizontal='right', vertical='top', wrapText=True)
        RC_ALIGN = Alignment(horizontal='right', vertical='center', wrapText=True)

        # Initial workbook 

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        ws.column_dimensions['A'].width = 15.46
        ws.column_dimensions['B'].width = 41.23
        ws.column_dimensions['C'].width = 10.31
        ws.column_dimensions['D'].width = 10.31
        ws.column_dimensions['E'].width = 41.23

        for i in (1, 2):
            row = ws.row_dimensions[i]
            row.fill = VIOLET_FILL
            row.border = OUTER_BORDER
            row.alignment = CC_ALIGN

        for i in (3, 4, 5):
            row = ws.row_dimensions[i]
            row.border = OUTER_BORDER
            row.alignment = CC_ALIGN

        row = ws.row_dimensions[6]
        row.fill = GREEN_FILL
        row.border = OUTER_BORDER
        row.alignment = CC_ALIGN

        for row in ws['A1:D5']:
            for cell in row:
                cell.border = OUTER_BORDER
                cell.alignment = CC_ALIGN

        values = ['Chip', 'Eng.', 'Date']

        for row, val in enumerate(values, start=2):
            cell = ws.cell(row, 1, val)
            cell.fill = ORANGE_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

        values = ['Number', 'FileName', 'Metion1', 'Metion2', 'PatternStatus']

        for row, val in enumerate(values, start=1):
            cell = ws.cell(row, 5, val)
            cell.fill = YELLOW_FILL
            cell.border = OUTER_BORDER
            cell.alignment = LC_ALIGN

        cell = ws.cell(6, 1, 'ADDR')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = CC_ALIGN

        cell = ws.cell(6, 2, 'Register')
        cell.fill = GREEN_FILL 
        cell.border = OUTER_BORDER
        cell.alignment = LC_ALIGN

        cell = ws.cell(6, 3, 'INI')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = RC_ALIGN

        cell = ws.cell(6, 4, 'Bits')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = RC_ALIGN

        cell = ws.cell(6, 5, 'Member')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = LC_ALIGN

        if is_init:
            cell = ws.cell(1, 6, 6)
            cell.fill = VIOLET_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

            cell = ws.cell(2, 6, 'PAT-1')
            cell.fill = VIOLET_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

        # Dump register

        row_st = row_ed = 7

        for addr in sorted(tuple(self.reg_table.keys())):
            reg_list = self.reg_table[addr]
            is_first_reg = True

            for reg in sorted(reg_list.regs, key=lambda reg: reg.lsb):
                cell_fill = YELLOW_FILL if is_first_reg else PatternFill()

                if is_first_reg:
                    is_first_reg = False
                    ws.cell(row_ed, 1, hex(addr))
                    ws.cell(row_ed, 2, reg_list.title)

                cell = ws.cell(row_ed, 1)
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = CT_ALIGN

                cell = ws.cell(row_ed, 2)
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = LT_ALIGN

                cell = ws.cell(row_ed, 3, hex(reg.init_val))
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = RT_ALIGN

                if reg.msb == reg.lsb:
                    cell = ws.cell(row_ed, 4, str(reg.msb))
                else:
                    cell = ws.cell(row_ed, 4, '_'.join([str(reg.msb), str(reg.lsb)]))

                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = RT_ALIGN

                cell_font = Font() if reg.is_access else GREY_FONT

                if reg.name == 'RESERVED':
                    cell = ws.cell(row_ed, 5, reg.name.lower())
                else:
                    toks = [] if reg.comment is None else reg.comment.split(',')
                    members = [reg.name] + [tok.strip() for tok in toks]
                    cell = ws.cell(row_ed, 5, '\n'.join(members))

                cell.font = cell_font
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = LT_ALIGN

                if is_init:
                    cell = ws.cell(row_ed, 6, f"{reg.init_val:X}")
                    cell.font = cell_font
                    cell.fill = cell_fill
                    cell.border = OUTER_BORDER
                    cell.alignment = CC_ALIGN

                row = ws.row_dimensions[row_ed]
                row.font = cell_font
                row.fill = cell_fill
                row.border = OUTER_BORDER
                row.alignment = CC_ALIGN
                row_ed += 1

        row = ws.row_dimensions[row_ed]
        row.fill = GREY_FILL
        row.border = OUTER_BORDER
        row.alignment = CC_ALIGN

        cell = ws.cell(row_ed, 1, 'none')
        cell.fill = GREY_FILL
        cell.border = OUTER_BORDER
        cell.alignment = CC_ALIGN
        row_ed += 1

        wb.save("table_dump.xlsx")
        wb.close()
    #}}}

    def str2int(self, str_: str) -> int:
        """Convert string to integer (with HEX check)"""  #{{{
        if str_.startswith('0x') or str_.startswith('0X') :
            return int(str_, 16)
        else:
            return int(str_)
    #}}}

    def access_check(self, str_: str) -> bool:
        """Syntax check for access flag"""  #{{{
        str_ = str_.lower()

        if str_ == 'y':
            return True
        elif str_ == 'n':
            return False
        else:
            raise SyntaxError("access flga must be 'y' or 'n'")
    #}}}

    def show_reg_table(self, msg: str):
        """Show register table"""  #{{{
        print(msg)
        reg_cnt = 0

        for addr, reg_list in self.reg_table.items():
            print(f"Addr: {addr:#06x} Title: {reg_list.title}")
            for reg in reg_list.regs:
                print("  Reg:      {}, {}, {}, {}, {}".format(
                    reg.name, reg.msb, reg.lsb, reg.is_access, reg.init_val))
                print("  Comment:  {}".format(reg.comment))
                print("  RowIndex: {}\n".format(reg.row_idx))
                reg_cnt += 1

        print(f"register count: {reg_cnt}\n")
    #}}}

    def show_ini_table(self, msg: str):
        """Show ini table"""  #{{{
        print(msg)
        reg_cnt = 0

        for ini_grp in self.ini_table:
            print(f"Tag: {ini_grp.tag}")
            for reg in ini_grp.regs:
                blank = ' ' * (ini_grp.max_len - len(reg.name))
                print(f"  {reg.name}{blank} = {reg.init_val}")
                reg_cnt += 1

        print(f"\nregister count: {reg_cnt}\n")
    #}}}

### Main Function ###

def main(is_debug=False):
    """Main function"""  #{{{
    parser = argparse.ArgumentParser(
            formatter_class=argparse.RawTextHelpFormatter,
            description=textwrap.dedent('''
                Programming register table converter.
                '''))

    parser.add_argument('in_type', metavar='input_type', 
                                    help='input type (option: txt/xls)') 
    parser.add_argument('out_type', metavar='output_type', 
                                    help='output type (option: txt/xls)') 
    parser.add_argument('table_fp', metavar='table_in',
                                    help='reference table in') 

    parser.add_argument('-i', dest='is_init', action='store_true', 
                                help='create initial pattern')

    args = parser.parse_args()

    # Parser register table

    pat_list = RegisterTable(args.table_fp, args.in_type, is_debug)

    # Dump register table

    if args.out_type == 'txt':
        pat_list.txt_export(args.is_init)
    elif args.out_type == 'xls':
        pat_list.xls_export(args.is_init)
    else:
        raise ValueError(f"Unsupported output table type ({args.out_type})")
#}}}

if __name__ == '__main__':
    main(False)

# SPDX-License-Identifier: GPL-2.0-only
#
# Copyright (C) 2022 Yeh, Hsin-Hsien <yhh76227@gmail.com>
#
"""
Mask reserved registers in the excel-style table
"""
import argparse
import textwrap
import openpyxl
import sys
from openpyxl.styles import Font

from .utils.general import PROG_VERSION

### Function ###

def mask_rsv_reg(table_fp):
    """Hide reserved register"""  #{{{
    GREY_FONT = Font(color='ff808080')
    wb = openpyxl.load_workbook(table_fp)
    ws = wb.worksheets[0]
    addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]
    start_row = addr_col.index('ADDR') + 2
    end_row = addr_col.index('none') + 1
    for row_idx in range(start_row, end_row):
        if str((cell := ws.cell(row_idx, 5)).value).lower() == 'reserved':
            if cell.font.__getattr__('color').rgb.lower() != 'ff808080':
                ws.row_dimensions[row_idx].font = GREY_FONT
                for col_idx in range(5, ws.max_column+1):
                    if (cell := ws.cell(row_idx, col_idx)).value != None:
                        cell.font = GREY_FONT
    wb.save(table_fp)
    wb.close()
#}}}

### Main ###

def main():
    """Main function"""  #{{{
    parser = argparse.ArgumentParser(
            formatter_class=argparse.RawTextHelpFormatter,
            description=textwrap.dedent("""
                Mark reserved registers in the excel-style table.
                """))

    parser.add_argument('--version', action='version', version=PROG_VERSION)
    parser.add_argument('table_fp', metavar='table_in', help="reference table in")
    args = parser.parse_args()
    mask_rsv_reg(args.table_fp)
#}}}

if __name__ == '__main__':
    sys.exit(main())

# SPDX-License-Identifier: GPL-2.0-only
#
# Copyright (C) 2022 Yeh, Hsin-Hsien <yhh76227@gmail.com>
#
"""
Programming Register Parser
"""
import argparse
import copy
import os
import pickle
import shutil
import sys
import textwrap
from pathlib import Path
from typing import NamedTuple

import openpyxl

from .utils.general import PROG_VERSION
from .utils.general import str2int
from .utils.ref_table import ReferenceTable

### Class Definition ###

class Pat(NamedTuple):
##{{{
    name: str
    regs: dict
##}}}

class PatternList(ReferenceTable):
    """Programming pattern list"""  #{{{

    def __init__(self, table_fp: str, table_type: str, debug_mode: set=None):
    #{{{
        # pat_list = [pat1, pat2, ...]

        super().__init__(debug_mode)
        self.pat_list  = []

        if table_type == 'txt':
            self.txt_table_parser(table_fp)
        elif table_type == 'xlsx':
            self.xlsx_table_parser(table_fp)
        elif table_type == 'db':
            with open(table_fp, 'rb') as f:
                self.reg_table = pickle.load(f)
                self.ini_table = pickle.load(f)
        else:
            raise ValueError(f"unsupported register table type ({table_type})")
    #}}}

    def ini_parser(self, ini_fp: str, is_batch=False, start=0, end=0):
        """Pattern parser for INI format"""  #{{{
        cfg_fps = []
        if is_batch:
            with open(ini_fp, 'r') as f:
                tmp_fps = f.readlines()

            if start < 1:
                start = 1
            elif start > len(tmp_fps):
                start = len(tmp_fps)

            if end == 0 or end > len(tmp_fps):
                end = len(tmp_fps)
            elif end < start:
                end = start

            for i in range(start-1, end):
                cfg_fps.append(tmp_fps[i].strip())
        else:
            cfg_fps.append(ini_fp)

        for cfg_fp in cfg_fps:
            pat_regs = {}
            with open(cfg_fp, 'r') as f:
                line = f.readline()
                line_no = 1
                while line:
                    if line.startswith('['):
                        pass
                    elif line.startswith('#'):
                        pass
                    else:
                        toks = line.split()
                        try:
                            if len(toks) and toks[1] == '=':
                                pat_regs[toks[0].upper()] = toks[2]
                        except Exception as e:
                            print('-' * 60)
                            print("INIRegParseError: (line: {})".format(line_no))
                            print("syntax of register descriptor:")
                            print("  '<reg_name> = <value> [comment]'")
                            print('-' * 60)
                            raise e
                    line = f.readline()
                    line_no += 1

            if 'p' in self.debug_mode:
                print(f"=== INI READ ({cfg_fp}) ===")
                for item in pat_regs.items():
                    print(item)
                print()

            pat_name = os.path.basename(cfg_fp)
            pat_name = os.path.splitext(pat_name)[0]
            self.pat_list.append(Pat(pat_name, pat_regs))
    #}}}

    def hex_parser(self, hex_fp: str, is_batch=False, start=0, end=0):
        """Pattern parser for HEX format"""  #{{{
        cfg_fps = []
        if is_batch:
            with open(hex_fp, 'r') as f:
                tmp_fps = f.readlines()

            if start < 1:
                start = 1
            elif start > len(tmp_fps):
                start = len(tmp_fps)

            if end == 0 or end > len(tmp_fps):
                end = len(tmp_fps)
            elif end < start:
                end = start

            for i in range(start-1, end):
                cfg_fps.append(tmp_fps[i].rstrip())
        else:
            cfg_fps.append(hex_fp)

        for cfg_fp in cfg_fps:
            pat_regs = {}
            with open(cfg_fp, 'r') as f:
                line = f.readline()
                while line:
                    addr = int(line[0:4], 16)
                    val = int(line[4:12], 16)
                    if addr in self.reg_table:
                        table_regs = self.reg_table[addr].regs
                        for reg in table_regs:
                            mask = (1 << (reg.msb - reg.lsb + 1)) - 1
                            pat_regs[reg.name] = hex((val >> reg.lsb) & mask)
                    line = f.readline()

            if 'p' in self.debug_mode:
                print(f"=== HEX READ ({cfg_fp}) ===")
                for item in pat_regs.items():
                    print(item)
                print()

            pat_name = os.path.basename(cfg_fp)
            pat_name = os.path.splitext(pat_name)[0]
            self.pat_list.append(Pat(pat_name, pat_regs))
    #}}}

    def xlsx_parser(self, xlsx_fp: str, is_batch=False, start=0, end=0):
        """Pattern parser for INI format"""  #{{{
        wb = openpyxl.load_workbook(xlsx_fp, data_only=True)
        ws = wb.worksheets[0]

        if is_batch:
            if start < 6:
                start = 6
            elif start > ws.max_column:
                start = ws.max_column

            if end == 0 or end > ws.max_column:
                end = ws.max_column
            elif end < start:
                end = start
        elif start < 6:
            end = start = 6
        elif start > ws.max_column:
            end = start = ws.max_column
        else:
            end = start

        addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]
        name_col = tuple(ws.iter_cols(5, 5, None, None, True))[0]
        row_st = addr_col.index('ADDR') + 1
        row_ed = addr_col.index('none')

        for j in range(start, end+1):
            val_col = tuple(ws.iter_cols(j, j, None, None, True))[0]
            pat_name = str(val_col[1]).strip()
            if pat_name == 'None' or pat_name == '':
                continue

            try:
                if ws.cell(2, j).font.__getattr__('color').rgb.lower() == 'ff808080':
                    continue
            except Exception:
                pass

            pat_regs = {}
            for i in range(row_st, row_ed):
                reg_name = name_col[i].split('\n')[0]
                reg_name = reg_name.strip().upper()
                if reg_name != 'RESERVED':
                    try:
                        pat_regs[reg_name] = '0x' + str(val_col[i])
                    except Exception as e:
                        print('-' * 60)
                        print("ExcelRegParseError: (row: {})".format(i+1))
                        print('-' * 60)
                        raise e

            if 'p' in self.debug_mode:
                print(f"=== XLS READ ({pat_name}) ===")
                for item in pat_regs.items():
                    print(item)
                print()

            self.pat_list.append(Pat(pat_name, pat_regs))

        wb.close()
    #}}}

    def ini_dump(self, pat_dir, pat_name=None, pat_ext=None, is_force=False, info_dump=True):
        """Dump pattern with ini format""" #{{{
        if not pat_ext:
            pat_ext = '.ini'

        pat_cnt = 0
        pat_ignore = 0
        is_batch = len(self.pat_list) > 1

        for pat in self.pat_list:
            if pat_name:
                pname = pat_name + str(pat_cnt) if is_batch else pat_name
            else:
                pname = pat.name

            pat_path = pat_dir / (pname + pat_ext)

            if pat_path.exists() and not is_force:
                if input(f"{pname+pat_ext} existed, overwrite? (y/n) ").lower() != 'y':
                    print('Ignore')
                    pat_cnt += 1
                    pat_ignore += 1
                    continue

            with open(pat_path, 'w') as f:
                is_first_tag = True
                for ini_grp in self.ini_table:
                    if ini_grp.tag is not None:
                        if is_first_tag:
                            is_first_tag = False
                        else:
                            f.write("\n")
                        f.write(f"[{ini_grp.tag}]\n")

                    for reg in ini_grp.regs:
                        if reg.name == '===':
                            f.write("\n")
                            continue

                        if reg.is_access:
                            try:
                                if reg.name in pat.regs:
                                    if reg.type == 'str':
                                        reg_val = pat.regs[reg.name]
                                    elif reg.type == 'float':
                                        reg_val = float(pat.regs[reg.name])
                                    else:
                                        reg_bits = reg.msb - reg.lsb + 1
                                        reg_val = str2int(pat.regs[reg.name], reg.is_signed, reg_bits)
                                else:
                                    print(f"[Warning] '{reg.name.lower()}' is not found in pattern '{pat.name}', use default value.")
                                    reg_val = reg.init_val
                            except Exception as e:
                                print('-' * 60)
                                print("RegisterValueError:")
                                print("pattern:  {}".format(pat.name))
                                print("register: {}".format(reg.name))
                                print('-' * 60)
                                raise SyntaxError("RegisterValueError") 

                            if reg.name in self.hex_out:
                                mask = (1 << (reg.msb - reg.lsb + 1)) - 1
                                if -65536 <= reg_val <= 65535:
                                    reg_val &= 0xffff
                                    lens = ini_grp.max_len + 12
                                    f.write(f"{reg.name.lower()} = {reg_val:#06x}".ljust(lens))
                                else:
                                    reg_val &= 0xffffffff
                                    lens = ini_grp.max_len + 16
                                    f.write(f"{reg.name.lower()} = {reg_val:#010x}".ljust(lens))
                            else:
                                lens = ini_grp.max_len + 11
                                f.write(f"{reg.name.lower()} = {reg_val}".ljust(lens))

                            if reg.comment is not None:
                                f.write(f" # {reg.comment}\n")
                            else:
                                f.write("\n")
            pat_cnt += 1

        if info_dump:
            print()
            print(f"=== Number of pattern generated: {pat_cnt - pat_ignore}")
            print(f"=== Number of pattern ignored:   {pat_ignore}")
            print()
    #}}}

    def hex_dump(self, pat_dir, pat_name=None, pat_ext=None, is_force=False, info_dump=True):
        """Dump pattern with hex format"""  #{{{
        if not pat_ext:
            pat_ext = '.pat'

        pat_cnt = 0
        pat_ignore = 0
        is_batch = len(self.pat_list) > 1

        for pat in self.pat_list:
            if pat_name:
                pname = pat_name + str(pat_cnt) if is_batch else pat_name
            else:
                pname = pat.name

            pat_path = pat_dir / (pname + pat_ext)

            if pat_path.exists() and not is_force:
                if input(f"{pname+pat_ext} existed, overwrite? (y/n) ").lower() != 'y':
                    print('Ignore')
                    pat_cnt += 1
                    pat_ignore += 1
                    continue

            with open(pat_path, 'w') as f:
                for addr in range(0, max(self.reg_table.keys()) + 4, 4):
                    word_val = 0
                    try:
                        for reg in self.reg_table[addr].regs:
                            bits = reg.msb - reg.lsb + 1
                            mask = (1 << bits) - 1
                            is_reg_exist = reg.name in pat.regs

                            if reg.is_access:
                                if is_reg_exist:
                                    try:
                                        reg_val = str2int(pat.regs[reg.name], reg.is_signed, bits)
                                    except Exception as e:
                                        print('-' * 60)
                                        print("RegisterValueError:")
                                        print("pattern:  {}".format(pat.name))
                                        print("register: {}".format(reg.name))
                                        print('-' * 60)
                                        raise SyntaxError("RegisterValueError") 
                                else:
                                    print(f"[Warning] '{reg.name.lower()}' is not found in pattern '{pat.name}', use default value.")
                                    reg_val = reg.init_val
                            else:
                                reg_val = reg.init_val

                            word_val += (reg_val & mask) << reg.lsb

                        f.write("{:04x}{:08x}\n".format(addr, word_val))
                    except Exception as e:
                        if str(e) == 'RegisterValueError':
                            raise e
                        f.write("{:04x}{:08x}\n".format(addr, 0))
            pat_cnt += 1

        if info_dump:
            print()
            print(f"=== Number of pattern generated: {pat_cnt - pat_ignore}")
            print(f"=== Number of pattern ignored:   {pat_ignore}")
            print()
    #}}}

    def xlsx_dump(self, ref_fp : str, pat_dir, pat_name=None, is_force=False, is_init=False, info_dump=True):
        """Dump pattern with excel format""" #{{{
        pname = pat_name if pat_name else 'register'
        pat_path = pat_dir / (pname + '.xlsx')

        if pat_path.exists() and not is_force:
            if input(f"{pname+'.xlsx'} existed, overwrite? (y/n) ").lower() != 'y':
                print('Terminal')
                exit(0)

        pat_cnt = 0
        wb = openpyxl.load_workbook(ref_fp)
        ws = wb.worksheets[0]

        if is_init and ws.max_column >= 6:
            ws.delete_cols(6, ws.max_column)

        pat_idx = ws.max_column + 1
        addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]
        row_st = addr_col.index('ADDR') + 2
        row_ed = addr_col.index('none') + 1

        for i in range(row_st, row_ed):
            str_ = ws.cell(i, 5).value
            reg_name = str_.split('\n')[0].strip()
            if reg_name.upper() == 'RESERVED':
                row = ws.row_dimensions[i]
                cell = ws.cell(i, pat_idx, 0)
                cell.font = copy.copy(row.font)
                cell.fill = copy.copy(row.fill)
                cell.border = copy.copy(row.border)
                cell.alignment = copy.copy(row.alignment)

        for pat in self.pat_list:
            for reg_list in self.reg_table.values():
                for reg in reg_list.regs:
                    bits = reg.msb - reg.lsb + 1
                    mask = (1 << bits) - 1

                    if not reg.is_access:
                        reg_val = 0
                    elif reg.name not in pat.regs:
                        print(f"[Warning] '{reg.name}' is not found in pattern '{pat.name}', use default value.")
                        reg_val = reg.init_val
                    else:
                        try:
                            reg_val = str2int(pat.regs[reg.name], reg.is_signed, bits)
                        except Exception as e:
                            print('-' * 60)
                            print("RegisterValueError:")
                            print("pattern:  {}".format(pat.name))
                            print("register: {}".format(reg.name))
                            print('-' * 60)
                            raise SyntaxError("RegisterValueError") 

                    row = ws.row_dimensions[reg.row_idx]
                    cell = ws.cell(reg.row_idx, pat_idx, hex(reg_val & mask).upper()[2:])
                    cell.font = copy.copy(row.font)
                    cell.fill = copy.copy(row.fill)
                    cell.border = copy.copy(row.border)
                    cell.alignment = copy.copy(row.alignment)

            row = ws.row_dimensions[1]
            cell = ws.cell(1, pat_idx, pat_idx)
            cell.font = copy.copy(row.font)
            cell.fill = copy.copy(row.fill)
            cell.border = copy.copy(row.border)
            cell.alignment = copy.copy(row.alignment)

            row = ws.row_dimensions[2]
            cell = ws.cell(2, pat_idx, pat.name)
            cell.font = copy.copy(row.font)
            cell.fill = copy.copy(row.fill)
            cell.border = copy.copy(row.border)
            cell.alignment = copy.copy(row.alignment)

            pat_idx += 1
            pat_cnt += 1

        wb.save(pat_path)
        wb.close()

        if info_dump:
            print(f"\n=== Number of pattern generated: {pat_cnt}\n")
    #}}}

    def export_table_db(self, db_fp):
        """Export reference table dtabase (pickle type)"""  #{{{
        with open(db_fp, 'wb') as f:
            pickle.dump(self.reg_table, f)
            pickle.dump(self.ini_table, f)
    #}}}
#}}}

### Main Function ###

def main():
    """Main function""" #{{{
    parser = argparse.ArgumentParser(
            formatter_class=argparse.RawTextHelpFormatter,
            description=textwrap.dedent("""
                Programming Register convertor.

                Convert register setting between ini/hex/excel format use the reference table.
                Output patterns will be exported to the new directory 'progp_out' by default.

                Single-source Examples:

                    @: %(prog)s -t table.txt ini hex reg.ini
                    
                        Convert a setting from ini to hex by text-style reference table.

                    @: %(prog)s -x table.xlsx ini xlsx reg.ini

                        Copy excel-style reference table and append a converted setting.

                    @: %(prog)s -X table.xlsx ini xlsx reg.ini

                        Create an empty excel reference table and append a converted setting.
                        
                Multi-source Examples:

                    @: %(prog)s -t table.txt ini hex <src_list_path> -b

                        Batch mode, convert all settings in the list.

                    @: %(prog)s -t table.txt ini hex <src_list_path> -b -s 2 -e 5

                        Batch mode, convert settings from the 2nd row to the 5th row in the list.

                    @: %(prog)s -t table.txt xlsx ini <excel_pat_list> -b -s 6 -e 8

                        Batch mode, convert settings from the 6th column to 8th column in the excel table.

                Convert to ini/hex with any format of reference table is permitted, but convert
                to excel format by excel-style reference table is necessary.
                """))

    parser.add_argument('in_fmt', metavar='format_in', choices=['ini', 'hex', 'xlsx'],
                                    help="input format (choices: ini/hex/xlsx)") 
    parser.add_argument('out_fmt', metavar='format_out', choices=['ini', 'hex', 'xlsx'], 
                                    help="output format (choices: ini/hex/xlsx)") 
    parser.add_argument('pat_in_fp', metavar='pattern_in',
                                    help="input pattern path") 

    parser.add_argument('--version', action='version', version=PROG_VERSION)

    table_gparser = parser.add_mutually_exclusive_group(required=True)
    table_gparser.add_argument('-t', dest='txt_table_fp', metavar='<path>',
                                        help="use text-style reference table")
    table_gparser.add_argument('-x', dest='xlsx_table_fp', metavar='<path>',
                                        help=textwrap.dedent("""\
                                        use excel-style reference table (current table append)"""))
    table_gparser.add_argument('-X', dest='xlsx_table_fp2', metavar='<path>',
                                        help=textwrap.dedent("""\
                                        use excel-style reference table (new table create)"""))
    # table_gparser.add_argument('-d', dest='database_fp', metavar='<path>',
    #                                     help=textwrap.dedent("""\
    #                                     use pre-parsed reference table database (pickle type)"""))

    # parser.add_argument('-p', dest='pickle_out_fp', metavar='<path>',
    #                             help=textwrap.dedent("""\
    #                             export reference table database (pickle type)"""))

    parser.add_argument('-b', dest='is_batch', action='store_true', 
                                help="enable batch mode")
    parser.add_argument('-s', dest='start_id', metavar='<id>', type=int, default=0,
                                help="start pattern index")
    parser.add_argument('-e', dest='end_id', metavar='<id>', type=int, default=0,
                                help="end pattern index")

    parser.add_argument('-f', dest='is_force', action='store_true', 
                                    help="force write with custom pattern dump")
    parser.add_argument('--dir', dest='cus_dir', metavar='<path>',
                                    help="custom dump directory")
    parser.add_argument('--pat', dest='cus_pat', metavar='<path>',
                                    help="custom dump pattern name")
    parser.add_argument('--ext', dest='cus_ext', metavar='<ext>',
                                    help="custom dump file extension (excel ignore)")

    args, args_dbg = parser.parse_known_args()

    parser_dbg = argparse.ArgumentParser()
    parser_dbg.add_argument('--dbg', dest='debug_mode', metavar='<pattern>',
                                        help="debug mode (tag: t/p)")

    args_dbg = parser_dbg.parse_known_args(args_dbg)[0]

    debug_mode = set()
    try:
        for i in range(len(args_dbg.debug_mode)):
            debug_mode.add(args_dbg.debug_mode[i])
    except Exception:
        pass

    ## Parser register table

    if args.txt_table_fp:
        pat_list = PatternList(args.txt_table_fp, 'txt', debug_mode)
    elif args.xlsx_table_fp:
        pat_list = PatternList(args.xlsx_table_fp, 'xlsx', debug_mode)
    elif args.xlsx_table_fp2:
        pat_list = PatternList(args.xlsx_table_fp2, 'xlsx', debug_mode)
    # else:
    #     pat_list = PatternList(args.database_fp, 'db', debug_mode)

    ## Only dump reference table database

    # if args.pickle_out_fp:
    #     pat_list.export_table_db(args.pickle_out_fp)
    #     return 0

    ## Parse input pattern

    if args.in_fmt == 'ini':
        pat_list.ini_parser(args.pat_in_fp, args.is_batch, args.start_id, args.end_id) 
    elif args.in_fmt == 'hex':
        pat_list.hex_parser(args.pat_in_fp, args.is_batch, args.start_id, args.end_id)
    else:
        pat_list.xlsx_parser(args.pat_in_fp, args.is_batch, args.start_id, args.end_id)

    ## Dump pattern

    if not args.cus_dir:
        pat_dir = Path('progp_out')
        if pat_dir.exists():
            shutil.rmtree(pat_dir) if pat_dir.is_dir() else pat_dir.unlink()
        pat_dir.mkdir()
    else:
        if not args.is_batch or args.out_fmt == 'xlsx':
            if (pat_dir := Path(args.cus_dir)).resolve() != Path().resolve():
                if pat_dir.exists():
                    if not args.is_force and input("output directory existed, overwrite? (y/n) ").lower() != 'y':
                        print('Terminated')
                        exit(0)
                    shutil.rmtree(pat_dir) if pat_dir.is_dir() else pat_dir.unlink()
                pat_dir.mkdir()
        else:
            if (pat_dir := Path(args.cus_dir)).resolve() == Path().resolve():
                print("[Error] custom dump directory can't be '.' in the batch mode.")
                exit(1)
            if pat_dir.exists():
                if not args.is_force and input("output directory existed, overwrite? (y/n) ").lower() != 'y':
                    print('Terminated')
                    exit(0)
                shutil.rmtree(pat_dir) if pat_dir.is_dir() else pat_dir.unlink()
            pat_dir.mkdir()

    pat_name = args.cus_pat if args.cus_pat else None

    try:
        pat_ext = '.' + args.cus_ext.split('.')[-1]
    except Exception:
        pat_ext = None

    if args.out_fmt == 'ini':
        pat_list.ini_dump(pat_dir, pat_name, pat_ext, args.is_force)
    elif args.out_fmt == 'hex':
        pat_list.hex_dump(pat_dir, pat_name, pat_ext, args.is_force)
    else:
        is_init = True if args.xlsx_table_fp2 else False

        if args.xlsx_table_fp:
            pat_list.xlsx_dump(args.xlsx_table_fp, pat_dir, pat_name, args.is_force, is_init)
        elif args.xlsx_table_fp2:
            pat_list.xlsx_dump(args.xlsx_table_fp2, pat_dir, pat_name, args.is_force, is_init)
        else:
            raise TypeError("need an excel register table when output excel file")
#}}}

if __name__ == '__main__':
    sys.exit(main())

